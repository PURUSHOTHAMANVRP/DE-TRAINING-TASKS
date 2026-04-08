import re
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

INPUT_FILE = "Delta3_Apr.xlsx"
OUTPUT_FILE = "Delta3_Ouptut.xlsx"



ROW_REVENUE = "Revenue"
ROW_REVENUE_PCT = "Revenue %"
ROW_SAL_ALLOC = "Total Salary Allocation for project"
ROW_SAL_ALLOC_PCT = "Total Salary Allocation %"
ROW_WORKFORCE = "Total Workforce Cost (CoreEmp + TL + Cons + Inc)"
ROW_TECH_SALARY_PCT = "Tech Salary Cost %"

def to_number(x):
   
   
    if pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace("£", "").replace(",", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    return float(s) if s else None


def read_sheet_metrics_as_dict(excel_path, sheet_name):
    
    

    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, engine="openpyxl")
    df = df.iloc[:, :2]                  
    df.columns = ["metric", "value"]
    df["metric"] = df["metric"].astype(str).str.strip()
    df["value"] = df["value"].apply(to_number)

    

    df = df[df["metric"].notna() & (df["metric"] != "nan") & (df["metric"] != "")]

    return dict(zip(df["metric"], df["value"]))


def find_cell_exact(ws, text):
    
    target = text.strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() == target:
                return cell.row, cell.column
    return None, None

def find_month_column(ws, month_str):
   
    for row in ws.iter_rows(min_row=1, max_row=25):
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, datetime):
                v = cell.value.strftime("%m/%d/%Y")
            else:
                v = str(cell.value).strip()
            if v == month_str:
                return cell.column
    return None

def main():
   
    xls = pd.ExcelFile(INPUT_FILE, engine="openpyxl")
    input_sheets = xls.sheet_names
    print("Input sheets:", input_sheets)

    
    sheet_to_month = {
        input_sheets[0]: "04/01/2025",
        "May": "05/01/2025",
        "June": "06/01/2025",
    }

    
    wb_out = load_workbook(OUTPUT_FILE)
    ws_out = wb_out[wb_out.sheetnames[0]]

    
    needed_rows = [
        ROW_REVENUE, ROW_REVENUE_PCT, ROW_SAL_ALLOC, ROW_SAL_ALLOC_PCT,
        ROW_WORKFORCE, ROW_TECH_SALARY_PCT
    ]
    row_num = {}
    for label in needed_rows:
        r, _ = find_cell_exact(ws_out, label)
        if r is None:
            raise ValueError(f"Row label not found in output: '{label}'")
        row_num[label] = r

    
    

    for sheet_name, month_header in sheet_to_month.items():
        if sheet_name not in input_sheets:
            print(f"Skipping missing input sheet: {sheet_name}")
            continue

        metrics = read_sheet_metrics_as_dict(INPUT_FILE, sheet_name)

        month_col = find_month_column(ws_out, month_header)
        if month_col is None:
            raise ValueError(f"Month header not found in output: '{month_header}'")

       
       
        revenue = metrics.get("Revenue")
        revenue_pct = metrics.get("Revenue %")
        sal_alloc = metrics.get("Total salary allocation for project")
        sal_alloc_pct = metrics.get("Total salary allocation %")

        if revenue is not None:
            ws_out.cell(row=row_num[ROW_REVENUE], column=month_col).value = revenue
        if revenue_pct is not None:
            ws_out.cell(row=row_num[ROW_REVENUE_PCT], column=month_col).value = revenue_pct
        if sal_alloc is not None:
            ws_out.cell(row=row_num[ROW_SAL_ALLOC], column=month_col).value = sal_alloc
        if sal_alloc_pct is not None:
            ws_out.cell(row=row_num[ROW_SAL_ALLOC_PCT], column=month_col).value = sal_alloc_pct

       
       
        core = metrics.get("Salary - Core employees") or 0
        tl = metrics.get("Salary - TL / Managers") or 0
        cons = metrics.get("Salary - Consultants") or 0
        inc = metrics.get("Performance payments - Incentive & Others") or 0

        workforce_cost = core + tl + cons + inc
        ws_out.cell(row=row_num[ROW_WORKFORCE], column=month_col).value = workforce_cost

       
        if revenue not in (None, 0):
            tech_pct = workforce_cost / revenue
            ws_out.cell(row=row_num[ROW_TECH_SALARY_PCT], column=month_col).value = tech_pct
        else:
            tech_pct = None

        print(f" {month_header} updated | Workforce={workforce_cost} | Tech%={tech_pct}")

    
    wb_out.save(OUTPUT_FILE)
    print(f"\n Updated in-place: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()